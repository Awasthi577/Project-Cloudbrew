#!/usr/bin/env python3
"""
Test CloudBrew's ability to handle all AWS resource types
Generates a report of which resources can be planned successfully
"""

import sys
import os
import subprocess
import json
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill

# Add CloudBrew to path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Common AWS resource types to test
# This is a sample - you can expand this list to include all 1580+ types
COMMON_AWS_RESOURCES = [
    # Compute
    "aws_instance",
    "aws_launch_template",
    "aws_autoscaling_group",
    
    # Storage
    "aws_s3_bucket",
    "aws_ebs_volume",
    "aws_efs_file_system",
    
    # Database
    "aws_dynamodb_table",
    "aws_db_instance",
    "aws_rds_cluster",
    
    # Networking
    "aws_vpc",
    "aws_subnet",
    "aws_security_group",
    "aws_lb",
    "aws_alb",
    "aws_route53_zone",
    
    # Serverless
    "aws_lambda_function",
    "aws_api_gateway_rest_api",
    "aws_sqs_queue",
    "aws_sns_topic",
    
    # Management
    "aws_iam_role",
    "aws_iam_policy",
    "aws_cloudwatch_alarm",
    
    # Containers
    "aws_eks_cluster",
    "aws_ecs_cluster",
    "aws_ecr_repository",
    
    # Developer Tools
    "aws_codebuild_project",
    "aws_codepipeline",
    "aws_codecommit_repository",
    
    # Analytics
    "aws_kinesis_stream",
    "aws_athena_database",
    "aws_glue_catalog_database",
    
    # Machine Learning
    "aws_sagemaker_notebook_instance",
    "aws_rekognition_collection",
    
    # Security
    "aws_kms_key",
    "aws_secretsmanager_secret",
    "aws_guardduty_detector",
    
    # Migration
    "aws_dms_replication_instance",
    "aws_sms_instance",
    
    # Media Services
    "aws_mediaconvert_queue",
    "aws_mediastore_container",
    
    # Robotics
    "aws_robomaker_robot_application",
    
    # Blockchain
    "aws_qldb_ledger",
    
    # Satellite
    "aws_groundstation_config",
    
    # Well-Architected
    "aws_wellarchitected_workload",
    
    # Additional common resources
    "aws_cloudfront_distribution",
    "aws_elasticache_cluster",
    "aws_elasticbeanstalk_application",
    "aws_stepfunctions_state_machine",
    "aws_waf_web_acl",
    "aws_wafv2_web_acl",
    "aws_xray_encryption_config",
    "aws_backup_plan",
    "aws_backup_vault",
    "aws_batch_compute_environment",
    "aws_budgets_budget",
    "aws_cognito_user_pool",
    "aws_connect_instance",
    "aws_datasync_task",
    "aws_dax_cluster",
    "aws_detective_graph",
    "aws_devicefarm_project",
    "aws_dlm_lifecycle_policy",
    "aws_docdb_cluster",
    "aws_drs_replication_configuration",
    "aws_dx_connection",
    "aws_ec2_client_vpn_endpoint",
    "aws_ec2_transit_gateway",
    "aws_ecrpublic_repository",
    "aws_ecs_task_definition",
    "aws_efs_access_point",
    "aws_eks_node_group",
    "aws_elasticache_replication_group",
    "aws_elasticsearch_domain",
    "aws_emr_cluster",
    "aws_fis_experiment_template",
    "aws_fms_policy",
    "aws_fsx_lustre_file_system",
    "aws_gamelift_game_session_queue",
    "aws_globalaccelerator_accelerator",
    "aws_glue_job",
    "aws_grafana_workspace",
    "aws_greengrass_group",
    "aws_guardduty_filter",
    "aws_healthlake_fhir_datastore",
    "aws_iam_instance_profile",
    "aws_iam_user",
    "aws_imagebuilder_image_pipeline",
    "aws_inspector_assessment_template",
    "aws_iot_thing",
    "aws_ivs_channel",
    "aws_kendra_index",
    "aws_keyspaces_table",
    "aws_lakeformation_data_lake_settings",
    "aws_lex_bot",
    "aws_licensemanager_license",
    "aws_lightsail_instance",
    "aws_location_place_index",
    "aws_macie2_account",
    "aws_managedblockchain_network",
    "aws_marketplace_catalog",
    "aws_mwaa_environment",
    "aws_neptune_cluster",
    "aws_networkfirewall_firewall",
    "aws_networkmanager_global_network",
    "aws_oam_link",
    "aws_opensearch_domain",
    "aws_opsworks_stack",
    "aws_organizations_account",
    "aws_panorama_package",
    "aws_personalize_dataset",
    "aws_pinpoint_app",
    "aws_prometheus_workspace",
    "aws_proton_service",
    "aws_qldb_stream",
    "aws_quicksight_data_source",
    "aws_ram_resource_share",
    "aws_rbin_rule",
    "aws_redshift_cluster",
    "aws_resiliencehub_app",
    "aws_resourceexplorer_index",
    "aws_robomaker_simulation_application",
    "aws_rolesanywhere_trust_anchor",
    "aws_route53_resolver_endpoint",
    "aws_s3_bucket_policy",
    "aws_s3control_bucket_policy",
    "aws_sagemaker_model",
    "aws_scheduler_schedule",
    "aws_schemas_schema",
    "aws_secretsmanager_secret_version",
    "aws_securityhub_account",
    "aws_servicecatalog_portfolio",
    "aws_ses_domain_identity",
    "aws_ses_email_identity",
    "aws_sfn_activity",
    "aws_shield_protection",
    "aws_signer_signing_profile",
    "aws_sns_platform_application",
    "aws_sqs_queue_policy",
    "aws_ssm_activation",
    "aws_ssm_document",
    "aws_ssm_maintenance_window",
    "aws_ssm_parameter",
    "aws_ssm_patch_baseline",
    "aws_ssmincidents_replication_set",
    "aws_storagegateway_gateway",
    "aws_synthetics_canary",
    "aws_timestreamwrite_database",
    "aws_transfer_server",
    "aws_trustedadvisor_checks",
    "aws_vpclattice_service",
    "aws_waf_rate_based_rule",
    "aws_wafregional_rate_based_rule",
    "aws_wafv2_ip_set",
    "aws_worklink_fleet",
    "aws_workspaces_directory",
    "aws_xray_sampling_rule"
]

def test_resource_type(resource_type: str) -> Dict[str, Any]:
    """Test a single resource type with CloudBrew"""
    result = {
        'resource_type': resource_type,
        'timestamp': datetime.now().isoformat(),
        'success': False,
        'error': '',
        'configuration': '',
        'plan_output': ''
    }
    
    try:
        # Build the command
        cmd = [
            sys.executable, '-m', 'LCF.cli',
            'intelligent-create',
            resource_type, 'test-resource',
            '--plan-only', '--yes'
        ]
        
        # Run the command
        process = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=60,  # 60 second timeout per resource
            cwd=os.path.dirname(os.path.abspath(__file__))
        )
        
        result['success'] = process.returncode == 0
        result['error'] = process.stderr if process.stderr else ''
        result['configuration'] = process.stdout if process.stdout else ''
        
        # Try to parse any JSON output
        try:
            if process.stdout:
                output_lines = process.stdout.split('\n')
                for line in output_lines:
                    if line.strip().startswith('{'):
                        result['plan_output'] = json.loads(line.strip())
                        break
        except json.JSONDecodeError:
            pass
            
    except subprocess.TimeoutExpired:
        result['error'] = 'Command timed out after 60 seconds'
    except Exception as e:
        result['error'] = f'Exception: {str(e)}'
    
    return result

def run_comprehensive_test() -> List[Dict[str, Any]]:
    """Run tests on all resource types"""
    results = []
    
    print(f"Starting CloudBrew AWS Resource Testing")
    print(f"Testing {len(COMMON_AWS_RESOURCES)} resource types...")
    print("=" * 80)
    
    for i, resource_type in enumerate(COMMON_AWS_RESOURCES):
        print(f"[{i+1}/{len(COMMON_AWS_RESOURCES)}] Testing {resource_type}...")
        
        try:
            result = test_resource_type(resource_type)
            results.append(result)
            
            if result['success']:
                print(f"  ✅ SUCCESS: {resource_type}")
            else:
                print(f"  ❌ FAILED: {resource_type}")
                if 'Command timed out' in result['error']:
                    print(f"     (Timeout - OpenTofu validation took too long)")
                else:
                    print(f"     Error: {result['error'][:100]}...")
                    
        except Exception as e:
            print(f"  ❌ EXCEPTION: {resource_type} - {str(e)}")
            results.append({
                'resource_type': resource_type,
                'timestamp': datetime.now().isoformat(),
                'success': False,
                'error': f'Test exception: {str(e)}',
                'configuration': '',
                'plan_output': ''
            })
    
    return results

def save_results_to_excel(results: List[Dict[str, Any]], filename: str = 'cloudbrew_resource_test_results.xlsx') -> None:
    """Save test results to Excel file"""
    print(f"\nSaving results to {filename}...")
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"
    
    # Add headers
    headers = [
        "Resource Type",
        "Success",
        "Error",
        "Configuration Generated",
        "Plan Output",
        "Timestamp"
    ]
    ws.append(headers)
    
    # Add data
    for result in results:
        row = [
            result['resource_type'],
            "YES" if result['success'] else "NO",
            result['error'][:100] if result['error'] else "None",
            "YES" if "Generated Configuration:" in result['configuration'] else "NO",
            "YES" if result['plan_output'] else "NO",
            result['timestamp']
        ]
        ws.append(row)
    
    # Add formatting
    success_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    failure_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        if row[1].value == "YES":  # Success column
            for cell in row:
                cell.fill = success_fill
        else:
            for cell in row:
                cell.fill = failure_fill
    
    # Add summary statistics
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Test Summary"])
    summary_ws.append(["Total Resources Tested", len(results)])
    
    success_count = sum(1 for r in results if r['success'])
    failure_count = len(results) - success_count
    
    summary_ws.append(["Successful", success_count])
    summary_ws.append(["Failed", failure_count])
    summary_ws.append(["Success Rate", f"{success_count/len(results)*100:.1f}%"])
    
    # Add resource type breakdown
    summary_ws.append(["", ""])
    summary_ws.append(["Resource Type Categories"])
    
    # Count by category (simplified)
    compute_resources = [r for r in results if 'instance' in r['resource_type'] or 'autoscaling' in r['resource_type']]
    storage_resources = [r for r in results if 's3' in r['resource_type'] or 'ebs' in r['resource_type'] or 'efs' in r['resource_type']]
    database_resources = [r for r in results if 'db' in r['resource_type'] or 'dynamo' in r['resource_type'] or 'rds' in r['resource_type']]
    
    summary_ws.append(["Compute Resources", len(compute_resources), sum(1 for r in compute_resources if r['success'])])
    summary_ws.append(["Storage Resources", len(storage_resources), sum(1 for r in storage_resources if r['success'])])
    summary_ws.append(["Database Resources", len(database_resources), sum(1 for r in database_resources if r['success'])])
    
    # Save file
    wb.save(filename)
    print(f"✅ Results saved to {filename}")

def main():
    """Main function"""
    print("🚀 CloudBrew AWS Resource Type Tester")
    print("=" * 80)
    print("This script tests CloudBrew's ability to handle various AWS resource types")
    print("Results will be saved to an Excel file for analysis")
    print()
    
    # Run tests
    results = run_comprehensive_test()
    
    # Save results
    save_results_to_excel(results)
    
    # Print summary
    success_count = sum(1 for r in results if r['success'])
    total_count = len(results)
    success_rate = success_count / total_count * 100
    
    print("\n" + "=" * 80)
    print("📊 TEST SUMMARY")
    print("=" * 80)
    print(f"Total Resources Tested: {total_count}")
    print(f"Successful: {success_count}")
    print(f"Failed: {total_count - success_count}")
    print(f"Success Rate: {success_rate:.1f}%")
    print("=" * 80)
    
    if success_rate > 80:
        print("🎉 EXCELLENT: CloudBrew handles most AWS resources well!")
    elif success_rate > 60:
        print("👍 GOOD: CloudBrew handles many AWS resources well.")
    elif success_rate > 40:
        print("⚠️ FAIR: CloudBrew handles some AWS resources well.")
    else:
        print("❌ NEEDS IMPROVEMENT: CloudBrew struggles with many AWS resources.")
    
    print("\n📁 Results saved to cloudbrew_resource_test_results.xlsx")
    print("💡 Use this data to identify which resource types need improvement")

if __name__ == "__main__":
    main()