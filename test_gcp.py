#!/usr/bin/env python3
"""
Test CloudBrew's ability to handle common GCP resource types
Generates a report of which resources can be planned successfully
"""

import sys
import os
import subprocess
import json
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill

# Add CloudBrew to path (same approach as your AWS script)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Common GCP resource types to test (sample; expand as needed)
COMMON_GCP_RESOURCES = [
    # Compute & Networking
    "google_compute_instance",
    "google_compute_disk",
    "google_compute_image",
    "google_compute_network",
    "google_compute_subnetwork",
    "google_compute_firewall",
    "google_compute_router",
    "google_compute_global_address",
    "google_compute_forwarding_rule",

    # Storage & CDN
    "google_storage_bucket",
    "google_storage_bucket_object",
    "google_compute_region_disk",
    "google_cloud_run_service",
    "google_compute_backend_service",
    "google_compute_url_map",
    "google_compute_target_http_proxy",
    "google_compute_target_https_proxy",
    "google_compute_ssl_certificate",
    "google_cloudcdn_origin",

    # Databases & Big Data
    "google_sql_database_instance",
    "google_sql_database",
    "google_sql_user",
    "google_bigquery_dataset",
    "google_bigquery_table",
    "google_spanner_instance",
    "google_spanner_database",
    "google_dataproc_cluster",
    "google_dataflow_job",  # note: may be dataflow_job in some providers

    # Kubernetes & Containers
    "google_container_cluster",   # GKE cluster
    "google_container_node_pool",
    "google_artifact_registry_repository",

    # Serverless & Messaging
    "google_cloudfunctions_function",
    "google_cloud_run_service",
    "google_pubsub_topic",
    "google_pubsub_subscription",
    "google_tasks_queue",

    # IAM & Security
    "google_service_account",
    "google_service_account_key",
    "google_project_iam_member",
    "google_kms_key_ring",
    "google_kms_crypto_key",
    "google_secret_manager_secret",
    "google_cloud_identity_group",

    # Monitoring & Logging
    "google_monitoring_alert_policy",
    "google_logging_project_sink",

    # ML & AI
    "google_ml_engine_model",
    "google_ai_platform_model",

    # Other managed services
    "google_compute_instance_group",
    "google_compute_region_instance_group_manager",
    "google_active_directory_domain",
    "google_dns_managed_zone",
    "google_dns_record_set",
    "google_service_usage_consumer_quota_override",
    "google_project_service",
    "google_sourcerepo_repository",
    "google_cloud_scheduler_job",
    "google_run_service_iam_member",
    "google_workflows_workflow"
]

def test_resource_type(resource_type: str, timeout_seconds: int = 60) -> Dict[str, Any]:
    """Test a single GCP resource type with CloudBrew (LCF.cli)"""
    result = {
        'resource_type': resource_type,
        'timestamp': datetime.now().isoformat(),
        'success': False,
        'error': '',
        'configuration': '',
        'plan_output': ''
    }

    try:
        cmd = [
            sys.executable, '-m', 'LCF.cli',
            'intelligent-create',
            resource_type, 'test-resource',
            '--plan-only', '--yes'
        ]

        process = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            cwd=os.path.dirname(os.path.abspath(__file__))
        )

        result['success'] = process.returncode == 0
        result['error'] = process.stderr if process.stderr else ''
        result['configuration'] = process.stdout if process.stdout else ''

        # attempt to find/parse JSON plan lines
        try:
            if process.stdout:
                for line in process.stdout.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    # Prefer lines that look like JSON objects/arrays
                    if line.startswith('{') or line.startswith('['):
                        try:
                            result['plan_output'] = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            # continue searching other lines
                            continue
        except Exception:
            # silent fallback; plan_output stays ''
            pass

    except subprocess.TimeoutExpired:
        result['error'] = f'Command timed out after {timeout_seconds} seconds'
    except Exception as e:
        result['error'] = f'Exception: {str(e)}'

    return result


def run_comprehensive_test(timeout_seconds: int = 60) -> List[Dict[str, Any]]:
    """Run tests on all GCP resource types in COMMON_GCP_RESOURCES"""
    results: List[Dict[str, Any]] = []

    print(f"Starting CloudBrew GCP Resource Testing")
    print(f"Testing {len(COMMON_GCP_RESOURCES)} resource types...")
    print("=" * 80)

    for i, resource_type in enumerate(COMMON_GCP_RESOURCES):
        print(f"[{i+1}/{len(COMMON_GCP_RESOURCES)}] Testing {resource_type}...")

        try:
            result = test_resource_type(resource_type, timeout_seconds=timeout_seconds)
            results.append(result)

            if result['success']:
                print(f"  ✅ SUCCESS: {resource_type}")
            else:
                print(f"  ❌ FAILED: {resource_type}")
                if 'timed out' in result['error'].lower():
                    print(f"     (Timeout - provider validation/plan took too long)")
                else:
                    err_preview = result['error'][:200] if result['error'] else '(no stderr)'
                    print(f"     Error: {err_preview}...")
        except Exception as e:
            print(f"  ❌ EXCEPTION: {resource_type} - {str(e)}")
            results.append({
                'resource_type': resource_type,
                'timestamp': datetime.now().isoformat(),
                'success': False,
                'error': f'Test exception: {str(e)}',
                'configuration': '',
                'plan_output': ''
            })

    return results


def save_results_to_excel(results: List[Dict[str, Any]], filename: str = 'cloudbrew_gcp_resource_test_results.xlsx') -> None:
    """Save test results to an Excel file"""
    print(f"\nSaving results to {filename}...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    headers = [
        "Resource Type",
        "Success",
        "Error",
        "Configuration Generated",
        "Plan Output",
        "Timestamp"
    ]
    ws.append(headers)

    for result in results:
        row = [
            result['resource_type'],
            "YES" if result['success'] else "NO",
            result['error'][:200] if result['error'] else "None",
            "YES" if "Generated Configuration:" in result['configuration'] or "generated configuration" in result['configuration'].lower() else "NO",
            "YES" if result['plan_output'] else "NO",
            result['timestamp']
        ]
        ws.append(row)

    # Formatting fills
    success_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    failure_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=len(results) + 1):
        if row[1].value == "YES":
            for cell in row:
                cell.fill = success_fill
        else:
            for cell in row:
                cell.fill = failure_fill

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Test Summary"])
    summary_ws.append(["Total Resources Tested", len(results)])

    success_count = sum(1 for r in results if r['success'])
    failure_count = len(results) - success_count

    summary_ws.append(["Successful", success_count])
    summary_ws.append(["Failed", failure_count])
    summary_ws.append(["Success Rate", f"{(success_count/len(results)*100) if len(results) else 0:.1f}%"])

    # Simple category breakdown (heuristic)
    compute_resources = [r for r in results if 'compute' in r['resource_type'] or 'instance' in r['resource_type'] or 'cluster' in r['resource_type']]
    storage_resources = [r for r in results if 'storage' in r['resource_type'] or 'bucket' in r['resource_type'] or 'disk' in r['resource_type']]
    database_resources = [r for r in results if 'sql' in r['resource_type'] or 'bigquery' in r['resource_type'] or 'spanner' in r['resource_type']]

    summary_ws.append(["", ""])
    summary_ws.append(["Resource Type Categories"])
    summary_ws.append(["Compute Resources", len(compute_resources), sum(1 for r in compute_resources if r['success'])])
    summary_ws.append(["Storage Resources", len(storage_resources), sum(1 for r in storage_resources if r['success'])])
    summary_ws.append(["Database Resources", len(database_resources), sum(1 for r in database_resources if r['success'])])

    wb.save(filename)
    print(f"✅ Results saved to {filename}")


def main():
    print("🚀 CloudBrew GCP Resource Type Tester")
    print("=" * 80)
    print("This script tests CloudBrew's ability to handle various GCP resource types")
    print("Results will be saved to an Excel file for analysis\n")

    # If you want a longer timeout for slow validations, change timeout_seconds here:
    timeout_seconds = 60

    results = run_comprehensive_test(timeout_seconds=timeout_seconds)
    save_results_to_excel(results)

    success_count = sum(1 for r in results if r['success'])
    total_count = len(results)
    success_rate = success_count / total_count * 100 if total_count else 0.0

    print("\n" + "=" * 80)
    print("📊 TEST SUMMARY")
    print("=" * 80)
    print(f"Total Resources Tested: {total_count}")
    print(f"Successful: {success_count}")
    print(f"Failed: {total_count - success_count}")
    print(f"Success Rate: {success_rate:.1f}%")
    print("=" * 80)

    if success_rate > 80:
        print("🎉 EXCELLENT: CloudBrew handles most GCP resources well!")
    elif success_rate > 60:
        print("👍 GOOD: CloudBrew handles many GCP resources well.")
    elif success_rate > 40:
        print("⚠️ FAIR: CloudBrew handles some GCP resources well.")
    else:
        print("❌ NEEDS IMPROVEMENT: CloudBrew struggles with many GCP resources.")

    print("\n📁 Results saved to cloudbrew_gcp_resource_test_results.xlsx")
    print("💡 Use this data to identify which resource types need improvement")


if __name__ == "__main__":
    main()
