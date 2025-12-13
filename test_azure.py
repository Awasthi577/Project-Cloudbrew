#!/usr/bin/env python3
"""
Test CloudBrew's ability to handle common Azure (azurerm) resource types
Generates a report of which resources can be planned successfully
Matches structure/format of your AWS/GCP scripts:
 - per-resource 60s timeout (adjustable)
 - captures stdout/stderr, attempts to parse JSON plan output
 - writes an Excel file with colored rows and a summary sheet
"""

import sys
import os
import subprocess
import json
from datetime import datetime
from typing import List, Dict, Any
import openpyxl
from openpyxl.styles import PatternFill

# Add CloudBrew to path (same approach as your other scripts)
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

# Common Azure resources to test (sample — expand as needed)
COMMON_AZURE_RESOURCES = [
    # Compute
    "azurerm_virtual_machine",
    "azurerm_virtual_machine_scale_set",
    "azurerm_availability_set",
    "azurerm_image",
    "azurerm_managed_disk",

    # Networking
    "azurerm_virtual_network",
    "azurerm_subnet",
    "azurerm_network_interface",
    "azurerm_network_security_group",
    "azurerm_public_ip",
    "azurerm_load_balancer",
    "azurerm_application_gateway",
    "azurerm_frontdoor",
    "azurerm_dns_zone",
    "azurerm_dns_a_record",

    # Storage & CDN
    "azurerm_storage_account",
    "azurerm_storage_container",
    "azurerm_storage_blob",
    "azurerm_cdn_profile",
    "azurerm_cdn_endpoint",
    "azurerm_fileshare",  # some providers use azurerm_storage_share

    # Databases
    "azurerm_sql_server",
    "azurerm_sql_database",
    "azurerm_mysql_server",
    "azurerm_postgresql_server",
    "azurerm_cosmosdb_account",
    "azurerm_mariadb_server",
    "azurerm_sql_elasticpool",

    # Kubernetes & Containers
    "azurerm_kubernetes_cluster",  # AKS
    "azurerm_container_registry",
    "azurerm_container_group",

    # Serverless & Messaging
    "azurerm_function_app",
    "azurerm_app_service_plan",
    "azurerm_eventhub_namespace",
    "azurerm_eventhub",
    "azurerm_servicebus_namespace",
    "azurerm_servicebus_queue",
    "azurerm_storage_queue",

    # Identity & Security
    "azurerm_role_definition",
    "azurerm_role_assignment",
    "azurerm_user_assigned_identity",
    "azurerm_key_vault",
    "azurerm_key_vault_secret",
    "azurerm_security_center_subscription_pricing",

    # Monitoring & DevOps
    "azurerm_monitor_metric_alert",
    "azurerm_log_analytics_workspace",
    "azurerm_application_insights",
    "azurerm_dev_test_lab",

    # Analytics & Big Data
    "azurerm_data_factory",
    "azurerm_synapse_workspace",
    "azurerm_synapse_sql_pool",
    "azurerm_data_lake_store",

    # AI & ML / Specialized
    "azurerm_machine_learning_workspace",
    "azurerm_cognitive_account",

    # Governance & Management
    "azurerm_policy_definition",
    "azurerm_policy_assignment",
    "azurerm_management_group",
    "azurerm_resource_group",
    "azurerm_subscription",

    # Backup / DR / Other
    "azurerm_recovery_services_vault",
    "azurerm_backup_policy_vm",
    "azurerm_site_recovery_replication_policy",
    "azurerm_iothub",
    "azurerm_media_services_account"
]

def test_resource_type(resource_type: str, timeout_seconds: int = 60) -> Dict[str, Any]:
    """Test a single Azure resource type with CloudBrew (LCF.cli)"""
    result = {
        'resource_type': resource_type,
        'timestamp': datetime.now().isoformat(),
        'success': False,
        'error': '',
        'configuration': '',
        'plan_output': ''
    }

    try:
        cmd = [
            sys.executable, '-m', 'LCF.cli',
            'intelligent-create',
            resource_type, 'test-resource',
            '--plan-only', '--yes'
        ]

        process = subprocess.run(
            cmd,
            capture_output=True,
            text=True,
            timeout=timeout_seconds,
            cwd=os.path.dirname(os.path.abspath(__file__))
        )

        result['success'] = process.returncode == 0
        result['error'] = process.stderr if process.stderr else ''
        result['configuration'] = process.stdout if process.stdout else ''

        # attempt to find/parse JSON plan lines (first JSON object/array encountered)
        try:
            if process.stdout:
                for line in process.stdout.splitlines():
                    line = line.strip()
                    if not line:
                        continue
                    if line.startswith('{') or line.startswith('['):
                        try:
                            result['plan_output'] = json.loads(line)
                            break
                        except json.JSONDecodeError:
                            continue
        except Exception:
            pass

    except subprocess.TimeoutExpired:
        result['error'] = f'Command timed out after {timeout_seconds} seconds'
    except Exception as e:
        result['error'] = f'Exception: {str(e)}'

    return result


def run_comprehensive_test(timeout_seconds: int = 60) -> List[Dict[str, Any]]:
    """Run tests on all Azure resource types"""
    results: List[Dict[str, Any]] = []

    print(f"Starting CloudBrew Azure Resource Testing")
    print(f"Testing {len(COMMON_AZURE_RESOURCES)} resource types...")
    print("=" * 80)

    for i, resource_type in enumerate(COMMON_AZURE_RESOURCES):
        print(f"[{i+1}/{len(COMMON_AZURE_RESOURCES)}] Testing {resource_type}...")

        try:
            result = test_resource_type(resource_type, timeout_seconds=timeout_seconds)
            results.append(result)

            if result['success']:
                print(f"  ✅ SUCCESS: {resource_type}")
            else:
                print(f"  ❌ FAILED: {resource_type}")
                if 'timed out' in result['error'].lower():
                    print(f"     (Timeout - provider validation/plan took too long)")
                else:
                    err_preview = result['error'][:200] if result['error'] else '(no stderr)'
                    print(f"     Error: {err_preview}...")
        except Exception as e:
            print(f"  ❌ EXCEPTION: {resource_type} - {str(e)}")
            results.append({
                'resource_type': resource_type,
                'timestamp': datetime.now().isoformat(),
                'success': False,
                'error': f'Test exception: {str(e)}',
                'configuration': '',
                'plan_output': ''
            })

    return results


def save_results_to_excel(results: List[Dict[str, Any]], filename: str = 'cloudbrew_azure_resource_test_results.xlsx') -> None:
    """Save test results to an Excel file"""
    print(f"\nSaving results to {filename}...")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Results"

    headers = [
        "Resource Type",
        "Success",
        "Error",
        "Configuration Generated",
        "Plan Output",
        "Timestamp"
    ]
    ws.append(headers)

    for result in results:
        row = [
            result['resource_type'],
            "YES" if result['success'] else "NO",
            result['error'][:200] if result['error'] else "None",
            "YES" if "Generated Configuration:" in result['configuration'] or "generated configuration" in result['configuration'].lower() else "NO",
            "YES" if result['plan_output'] else "NO",
            result['timestamp']
        ]
        ws.append(row)

    # Formatting fills
    success_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
    failure_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    for row in ws.iter_rows(min_row=2, max_row=len(results) + 1):
        if row[1].value == "YES":
            for cell in row:
                cell.fill = success_fill
        else:
            for cell in row:
                cell.fill = failure_fill

    # Summary sheet
    summary_ws = wb.create_sheet("Summary")
    summary_ws.append(["Test Summary"])
    summary_ws.append(["Total Resources Tested", len(results)])

    success_count = sum(1 for r in results if r['success'])
    failure_count = len(results) - success_count

    summary_ws.append(["Successful", success_count])
    summary_ws.append(["Failed", failure_count])
    summary_ws.append(["Success Rate", f"{(success_count/len(results)*100) if len(results) else 0:.1f}%"])

    # Simple category breakdown (heuristic)
    compute_resources = [r for r in results if 'virtual_machine' in r['resource_type'] or 'vm' in r['resource_type'] or 'kubernetes' in r['resource_type']]
    network_resources = [r for r in results if 'network' in r['resource_type'] or 'subnet' in r['resource_type'] or 'load_balancer' in r['resource_type']]
    storage_resources = [r for r in results if 'storage' in r['resource_type'] or 'disk' in r['resource_type'] or 'cdn' in r['resource_type']]

    summary_ws.append(["", ""])
    summary_ws.append(["Resource Type Categories"])
    summary_ws.append(["Compute Resources", len(compute_resources), sum(1 for r in compute_resources if r['success'])])
    summary_ws.append(["Network Resources", len(network_resources), sum(1 for r in network_resources if r['success'])])
    summary_ws.append(["Storage Resources", len(storage_resources), sum(1 for r in storage_resources if r['success'])])

    wb.save(filename)
    print(f"✅ Results saved to {filename}")


def main():
    print("🚀 CloudBrew Azure Resource Type Tester")
    print("=" * 80)
    print("This script tests CloudBrew's ability to handle various Azure (azurerm) resource types")
    print("Results will be saved to an Excel file for analysis\n")

    # Adjust timeout_seconds for slow provider validations (e.g., AKS)
    timeout_seconds = 60

    results = run_comprehensive_test(timeout_seconds=timeout_seconds)
    save_results_to_excel(results)

    success_count = sum(1 for r in results if r['success'])
    total_count = len(results)
    success_rate = success_count / total_count * 100 if total_count else 0.0

    print("\n" + "=" * 80)
    print("TEST SUMMARY")
    print("=" * 80)
    print(f"Total Resources Tested: {total_count}")
    print(f"Successful: {success_count}")
    print(f"Failed: {total_count - success_count}")
    print(f"Success Rate: {success_rate:.1f}%")
    print("=" * 80)

    if success_rate > 80:
        print("CHETAAH HI BOLE DEE")
        print("EXCELLENT: CloudBrew handles most Azure resources well!")
    elif success_rate > 60:
        print("GOOD: CloudBrew handles many Azure resources well.")
    elif success_rate > 40:
        print("FAIR: CloudBrew handles some Azure resources well.")
    else:
        print("NEEDS IMPROVEMENT: CloudBrew struggles with many Azure resources.")

    print("\nResults saved to cloudbrew_azure_resource_test_results.xlsx")
    print("Use this data to identify which resource types need improvement")


if __name__ == "__main__":
    main()
